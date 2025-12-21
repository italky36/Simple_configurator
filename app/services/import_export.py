import csv
import io
from typing import Dict, Iterable, List, Tuple

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .. import crud, models

FIELD_HEADERS: List[Tuple[str, str]] = [
    ("model", "Модель оборудования"),
    ("frame", "Каркас"),
    ("frame_color", "Цвет каркаса"),
    ("frame_design_color", "Цвет дизайна каркаса"),
    ("refrigerator", "Холодильник"),
    ("terminal", "Терминал"),
    ("price", "Цена"),
    ("ozon_link", "Ссылка на Озон"),
    ("graphic_link", "Ссылка на графику"),
    ("main_image", "Main image"),
    ("gallery_folder", "Gallery folder"),
    ("description", "Описание"),
]

ALIAS_MAP = {
    "модель": "model",
    "модель оборудования": "model",
    "каркас": "frame",
    "цвет каркаса": "frame_color",
    "цвет дизайна каркаса": "frame_design_color",
    "цвет дизайна": "frame_design_color",
    "цвет вставки": "frame_design_color",
    "холодильник": "refrigerator",
    "терминал": "terminal",
    "цена": "price",
    "ссылка на озон": "ozon_link",
    "ссылка на графику": "graphic_link",
    "main image": "main_image",
    "gallery folder": "gallery_folder",
    "описание": "description",
    "name": "name",
}


def normalize_key(key: str) -> str:
    return key.strip().lower().replace("-", " ").replace("_", " ")


def map_row_keys(row: Dict[str, str]) -> Dict[str, str]:
    mapped: Dict[str, str] = {}
    for key, value in row.items():
        normalized = normalize_key(key)
        target = ALIAS_MAP.get(normalized, normalized.replace(" ", "_"))
        mapped[target] = value
    return mapped


def _parse_price(value: str):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _prepare_machine_data(raw: Dict[str, str]) -> Dict[str, str]:
    data = map_row_keys(raw)
    if not data.get("name"):
        data["name"] = data.get("model") or "Coffee Machine"
    # frame_design_color используется только если есть ссылка на Озон
    if not data.get("ozon_link"):
        data["frame_design_color"] = None
    if "price" in data:
        data["price"] = _parse_price(data.get("price"))
    return {k: v for k, v in data.items() if v not in (None, "")}


def import_file(db: Session, file: UploadFile, update_existing: bool = True) -> Dict[str, int]:
    filename = file.filename or ""
    if filename.lower().endswith(".csv"):
        rows = _read_csv(file)
    elif filename.lower().endswith((".xlsx", ".xlsm")):
        rows = _read_excel(file)
    else:
        raise HTTPException(status_code=400, detail="Поддерживаются только CSV или XLSX")

    created = updated = 0
    for raw in rows:
        data = _prepare_machine_data(raw)
        if not data.get("model") and not data.get("name"):
            continue
        # Построение "сигнатуры" варианта: модель + каркас + цвет + холодильник + терминал
        signature = (
            data.get("model"),
            data.get("frame"),
            data.get("frame_color"),
            data.get("frame_design_color"),
            data.get("refrigerator"),
            data.get("terminal"),
        )
        if update_existing and signature[0]:
            existing = crud.get_coffee_machine_by_signature(db, *signature)
            if existing:
                crud.update_coffee_machine(db, existing.id, data)
                updated += 1
                continue
        crud.create_coffee_machine(db, data)
        created += 1
    return {"created": created, "updated": updated}


def export_csv(machines: Iterable[models.CoffeeMachine]) -> io.StringIO:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([header for _, header in FIELD_HEADERS])
    for machine in machines:
        writer.writerow([getattr(machine, field, "") or "" for field, _ in FIELD_HEADERS])
    buffer.seek(0)
    return buffer


def export_xlsx(machines: Iterable[models.CoffeeMachine]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "coffee_machines"
    headers = [header for _, header in FIELD_HEADERS]
    ws.append(headers)
    header_fill = "FFD3E0"
    for idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = cell.font.copy(bold=True)
        cell.fill = cell.fill.copy(fgColor=header_fill)
    for machine in machines:
        ws.append([getattr(machine, field, "") or "" for field, _ in FIELD_HEADERS])
    # Немного ширины по умолчанию
    widths = {
        "A": 20,  # model
        "B": 16,  # frame
        "C": 14,  # frame_color
        "D": 14,  # frame_design_color
        "E": 14,  # refrigerator
        "F": 14,  # terminal
        "G": 10,  # price
        "H": 40,  # ozon_link
        "I": 40,  # graphic_link
        "J": 40,  # main_image
        "K": 24,  # gallery_folder
    }
    for col, width in widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    return data


def _read_csv(file: UploadFile) -> List[Dict[str, str]]:
    file.file.seek(0)
    text = io.TextIOWrapper(file.file, encoding="utf-8-sig")
    reader = csv.DictReader(text)
    return [dict(row) for row in reader]


def _read_excel(file: UploadFile) -> List[Dict[str, str]]:
    file.file.seek(0)
    wb = load_workbook(file.file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({headers[i]: row[i] for i in range(len(headers))})
    return data_rows
